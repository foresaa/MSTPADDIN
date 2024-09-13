import win32com.client
from win32com.client import constants
import pandas as pd
from tkinter import Tk, filedialog
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Function to format data as a table and autoformat columns
def format_excel_sheet(sheet, table_name):
    # Get the dimensions of the data
    last_row = sheet.max_row
    last_col = sheet.max_column

    # Define the table range
    table_range = f"A1:{openpyxl.utils.get_column_letter(last_col)}{last_row}"

    # Create and style the table
    tab = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    # Auto-size columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid issues with non-string data
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

# Function to open a file dialog for selecting a Project file
def select_project_file():
    root = Tk()
    root.withdraw()  # Hide the main tkinter window
    project_file = filedialog.askopenfilename(title="Select Microsoft Project File",
                                              filetypes=[("Microsoft Project Files", "*.mpp")])
    return project_file

# Main function to run the extraction and formatting
def extract_msp_fields_to_excel():
    # Open file dialog to select a Microsoft Project file
    project_file = select_project_file()
    
    if not project_file:
        print("No file selected. Exiting.")
        return

    # Dispatch the MS Project Application
    project_app = win32com.client.Dispatch("MSProject.Application")
    project_app.Visible = True  # Ensure MS Project is visible

    # Open the selected project file
    project_app.FileOpen(project_file)

    # Access the active project
    active_project = project_app.ActiveProject

    # Ensure there's an active project loaded
    if active_project is None:
        print("No active project found.")
        return

    # Define the PjValueListItem constant for value retrieval
    pjValueListValue = constants.pjValueListValue

    # Define the default names for all task and resource custom fields
    default_field_names = {}

    # Populate default names for task fields (Text1-30, Number1-20, etc.)
    for i in range(1, 31):
        default_field_names[getattr(constants, f'pjCustomTaskText{i}')] = f"Text{i} (Task)"
    for i in range(1, 21):
        default_field_names[getattr(constants, f'pjCustomTaskNumber{i}')] = f"Number{i} (Task)"
        default_field_names[getattr(constants, f'pjCustomTaskFlag{i}')] = f"Flag{i} (Task)"
    for i in range(1, 11):
        default_field_names[getattr(constants, f'pjCustomTaskStart{i}')] = f"Start{i} (Task)"
        default_field_names[getattr(constants, f'pjCustomTaskFinish{i}')] = f"Finish{i} (Task)"
        default_field_names[getattr(constants, f'pjCustomTaskDuration{i}')] = f"Duration{i} (Task)"
        default_field_names[getattr(constants, f'pjCustomTaskCost{i}')] = f"Cost{i} (Task)"
        default_field_names[getattr(constants, f'pjCustomTaskDate{i}')] = f"Date{i} (Task)"
        default_field_names[getattr(constants, f'pjCustomTaskOutlineCode{i}')] = f"OutlineCode{i} (Task)"

    # Populate default names for resource fields (Text1-30, Number1-20, etc.)
    for i in range(1, 31):
        default_field_names[getattr(constants, f'pjCustomResourceText{i}')] = f"Text{i} (Resource)"
    for i in range(1, 21):
        default_field_names[getattr(constants, f'pjCustomResourceNumber{i}')] = f"Number{i} (Resource)"
        default_field_names[getattr(constants, f'pjCustomResourceFlag{i}')] = f"Flag{i} (Resource)"
    for i in range(1, 11):
        default_field_names[getattr(constants, f'pjCustomResourceStart{i}')] = f"Start{i} (Resource)"
        default_field_names[getattr(constants, f'pjCustomResourceFinish{i}')] = f"Finish{i} (Resource)"
        default_field_names[getattr(constants, f'pjCustomResourceDuration{i}')] = f"Duration{i} (Resource)"
        default_field_names[getattr(constants, f'pjCustomResourceCost{i}')] = f"Cost{i} (Resource)"
        default_field_names[getattr(constants, f'pjCustomResourceDate{i}')] = f"Date{i} (Resource)"
        default_field_names[getattr(constants, f'pjCustomResourceOutlineCode{i}')] = f"OutlineCode{i} (Resource)"

    # Combine task and resource field constants for all 9 types
    custom_field_constants = list(default_field_names.keys())

    # Create lists to store field information for tasks and resources
    fields_info = []

    # Iterate through all custom fields for tasks and resources
    for custom_field_constant in custom_field_constants:
        try:
            # Get the custom field name, fallback to default name if necessary
            custom_field_name = project_app.CustomFieldGetName(custom_field_constant)
            if not custom_field_name:
                custom_field_name = default_field_names.get(custom_field_constant, "Unnamed Field")

            # Get the formula for the custom field, if any
            custom_field_formula = project_app.CustomFieldGetFormula(custom_field_constant)

            # Get lookup list items for the custom field
            lookup_items = []
            index = 1
            while True:
                try:
                    # Retrieve each lookup list item by index
                    lookup_item = project_app.CustomFieldValueListGetItem(custom_field_constant, pjValueListValue, index)
                    if lookup_item:
                        lookup_items.append(lookup_item)
                    index += 1
                except Exception as e:
                    # Break when no more lookup items are found
                    if "1101" in str(e):
                        break
                    else:
                        print(f"Unexpected error retrieving lookup items for {custom_field_name}: {e}")
                        break

            # Store the field information (default name, custom name, formula, lookup items)
            fields_info.append({
                'Default Field Name': default_field_names[custom_field_constant],
                'Custom Name': custom_field_name if custom_field_name != default_field_names[custom_field_constant] else "None",
                'Formula': custom_field_formula if custom_field_formula else "None",
                'Lookup Items': ', '.join(lookup_items) if lookup_items else "None"
            })

        except Exception as e:
            print(f"Error processing field {custom_field_constant}: {e}")

    # Convert fields into a DataFrame and sort by default field name and task/resource type
    fields_df = pd.DataFrame(fields_info)
    fields_df = fields_df.sort_values(by=['Default Field Name'])

    # Separate tasks and resources into different sheets
    tasks_df = fields_df[fields_df['Default Field Name'].str.contains('(Task)')]
    resources_df = fields_df[fields_df['Default Field Name'].str.contains('(Resource)')]

    # Write the data to an Excel workbook with two sheets (Tasks, Resources)
    workbook = openpyxl.Workbook()

    # Tasks Sheet
    task_sheet = workbook.active
    task_sheet.title = "Tasks"
    for r in dataframe_to_rows(tasks_df, index=False, header=True):
        task_sheet.append(r)
    format_excel_sheet(task_sheet, "TasksTable")

    # Resources Sheet
    resource_sheet = workbook.create_sheet("Resources")
    for r in dataframe_to_rows(resources_df, index=False, header=True):
        resource_sheet.append(r)
    format_excel_sheet(resource_sheet, "ResourcesTable")

    # Save the workbook
    workbook.save('custom_fields_output.xlsx')
    print("Custom fields have been exported to 'custom_fields_output.xlsx'.")

    # Close the project file without saving
    project_app.FileClose(False)  # False = don't save changes
    project_app.Quit()

# Run the main extraction function
extract_msp_fields_to_excel()
